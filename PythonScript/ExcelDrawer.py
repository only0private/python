# -*- coding: utf-8 -*-
"""
Created on Sun Nov 12 04:21:23 2017

@author: hamaguchi
"""

from PIL import Image
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import PatternFill
import xlwt
#==============================================================================
# Excelで描こう
#==============================================================================
img = Image.open("aaa.png")
width, height = img.size    #横，縦

img_pixels = []
for y in range(height):
  for x in range(width):
    img_pixels.append(img.getpixel((x,y)))
img_pixels = np.array(img_pixels)

file = "picture.xls"
book = xlwt.Workbook()
ns1 = book.add_sheet('NewSheet1') 
#xlwt.add_palette_colour("custom_colour", 0x21)

i = 0

for y in range(height):
    for x in range(0):
        print(img_pixels[i][0])
        xlwt.add_palette_colour("custom_colour", 0x21)
        book.set_colour_RGB(0x21, img_pixels[i][0], img_pixels[i][1], img_pixels[i][2])
        style = xlwt.easyxf('pattern: pattern solid, fore_colour custom_colour')
        ns1.write(y, x, '', style)
        i += 1
        book.save(file)



